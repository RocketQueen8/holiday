# coding:utf-8
import xlrd
import psutil
import xlutils
import xlwt
import openpyxl
import os
import pandas
import sys
import datetime
import json
from multiprocessing import Process,Pool
from collections import namedtuple,deque,defaultdict,OrderedDict,ChainMap,Counter
import hashlib
from urllib import request
import requests



# c=datetime.datetime.now()
# print(c)
# d=datetime.datetime.date(c)
# print(d)

# path=os.getcwd()
# print(path)
# t=psutil.disk_usage(path).total/1024/1024/1024
# u=psutil.disk_usage(path).used/1024/1024/1024
# f=psutil.disk_usage(path).free/1024/1024/1024
# p=psutil.disk_usage(path).percent
# print(f)
# print(u)
# print(t)
# print(p)

# base_path = '2021-09-29 10_29_26-cashApiInterFace数据.xlsx'
# ob = openpyxl.load_workbook(filename=base_path)
# sn = ob.sheetnames
# sh = ob[sn[0]]
# li = []
# for i in sh[47]:
#     li.append(i.value)
# print(li)
# print(li[1])
# print(type(li[1]))
# print(li[1].strip("/"))
# print(li)
# path=os.getcwd()
# print(path)
# disk=psutil.disk_usage(path)
# f=disk.free/1024/1024/1024
# print(f)
# print(disk)



# class HandleExcel:
#     def load_excel(self):
#         '''
#         加载excel
#         '''
#         open_excel = openpyxl.load_workbook(base_path)#拿到excel的所有内容
#         return open_excel
#     def get_sheet_data(self,index=None):
#         '''
#         加载所有sheet的内容
#         '''
#         sheet_name = self.load_excel().sheetnames#拿到sheetnames的所有内容
#         if index == None:
#             index = 0
#         data = self.load_excel()[sheet_name[index]]
#         return data
#     def get_cell_value(self,row,cols):
#         '''
#         获取某一个单元格内容
#         '''
#         data = self.get_sheet_data().cell(row=row,column=cols)
#         return data
#     def get_rows(self):
#         row = self.get_sheet_data().max_row
#         return row
#     def get_rows_value(self,row):
#         '''
#         获取某一行的内容
#         '''
#         row_list = []
#         for i in self.get_sheet_data()[row]:
#             row_list.append(i.value)
#         return row_list
#
# if __name__ == '__main__':
#     handle = HandleExcel()
#     print(handle.get_rows_value(2))

# try:
#     print("try.....")
#     r=10/0
#     print("result:",r)
# except ZeroDivisionError as e:
#     print("except:",e)
# finally:
#     print("finally......")

# f=open("demo.py","r",encoding="UTF-8")
# n=f.read()
# print(n)
# c=f.close()
# print(c)
# with open("demo2.py","r",encoding="utf-8") as f:
#     r=f.read()
#     print(r)

# with open("demo3.py","a",encoding="utf-8") as f:
#     f.write("aaaaaaa")
# e=os.environ
# n=os.name
# u=os.uname
# print(e)
# print(n)
# print(u)

# g=os.getcwd()
# print(g)
# # j=os.path.join(g,"testdir")
# # print(j)
# # os.mkdir(j)
# # os.rmdir(j)
# os.rename("demo3.c","demo3.py")

# d=dict(name="jack",age=28,ismale=True,status=None,score=95.5)
# j=json.dumps(d)
# print(j)
# class Student:
#     def __init__(self,name,age,ismale,status,score):
#         self.name=name
#         self.age=age
#         self.ismale=ismale
#         self.status=status
#         self.score=score
# s=Student("jack",28,True,None,95.5)
#
# def studentdict(std):
#     return {
#         "name":std.name,
#         "age":std.age,
#         "ismale":std.ismale,
#         "status":std.status,
#         "score":std.score
#     }
# print(type(j))
# d=json.loads(j)
# print(d)
# print(type(d))
# dd=json.dumps(s,default=studentdict)
# print(dd)

# def run_proc(name):
#     print("Run child process %s (%s)...." % (name,os.getpid()))
#
# if __name__ == "__main__":
#     print("Parent process %s." % os.getpid())
#     p=Process(target=run_proc,args=("test",))
#     print("Child process will start.")
#     p.start()
#     p.join()
#     print("Child process end.")

# now=datetime.datetime.now()
# dt=datetime.datetime(2021,11,22,16,40)
# tt=dt.timestamp()
# nt=now.timestamp()
# n=datetime.datetime.fromtimestamp(nt)
# u=datetime.datetime.utcfromtimestamp(nt)
# print(now)
# print(dt)
# print(tt)
# print(nt)
# print(n)
# print(u)

# S=datetime.datetime.strptime("2011-12-05 15:30:45","%Y-%m-%d %H:%M:%S")
# print(S)
# print(type(S))
# d=S.strftime("%Y-%m-%d %H:%M:%S")
# print(d)
# print(type(d))

# now=datetime.datetime.now()
# print(now)
# d=datetime.datetime(2021,6,25,12,25,36)
# print(d)
# d1=now+datetime.timedelta(hours=10)
# print(d1)
# d2=now+datetime.timedelta(days=3)
# print(d2)
# d3=now+datetime.timedelta(hours=5,days=5)
# print(d3)
# s=now.timestamp()
# print(s)
# sd=datetime.datetime.fromtimestamp(s)
# print(sd)
# s=datetime.datetime.strptime("2013-2-25 12:25:12","%Y-%m-%d %H:%M:%S")
# print(s)
# st=now.strftime("%Y-%m-%d %H:%M:%S")
# st1=now.strftime("%Y-%m-%d")
# st2=now.strftime("%H:%M:%S")
# st3=now.strftime("%Y-%m")
# print(st)
# print(st1)
# print(st2)
# print(st3)

# point=namedtuple("point",["x","y"])
# p=point(1,2)
# print(p.x)
# print(p.y)
# lis=deque(['a',"b","c","d"])
# lis.append("x")
# lis.appendleft("y")
# print(lis)
# lis.pop()
# lis.popleft()
# print(lis)

# res={
#     "code":200,
#     "status":"succesful",
#     "content":"{"
#               "'name':'jack',"
#               "'age':28,"
#               "'gender':'male'"
#               "}"
# }
# j=json.dumps(res)
# print(j)
# md5=hashlib.md5()
# md5.update(j.encode("utf-8"))
# print(md5.hexdigest())

# with request.urlopen('https://www.baidu.com') as f:
#     data=f.read()
#     # print(data)
#     print("Status:",f.status,f.reason)
#     for k,v in f.getheaders():
#         print("%s:%s"%(k,v))
#     print("Data:","\n",data.decode("utf-8"))


# with requests.get("https://www.baidu.com") as res:
#     data=res.text
#     head=res.headers
#     print("Data:",data)
#     print(head)
#     print(res.cookies)
#     print(res.encoding)


# cpu=psutil.cpu_count()
# cpus=psutil.cpu_count(logical=False)
# print(cpu)
# print(cpus)





# {
#   "method": "POST",
#   "encryptData": "{\"mobile\":\"8127458457\",\"type\":\"text\"}",
#   "sourceChannel": "Organic",
#   "ip": "223.104.20.234",
#   "packageName": "com.tkt.ins",
#   "contentType": "application/x-www-form-urlencoded",
#   "version": "5.7.0"
# }



#!/usr/bin/env python
#encoding=‘utf-8'




































